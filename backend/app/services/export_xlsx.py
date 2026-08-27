"""
Excel export services.

Supported exports:

1. generate_registry_xlsx(questions)
   Legacy export of individual questions.

2. generate_variants_registry_xlsx(variants)
   Export of complete variants with bilingual questions.
"""

import io
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


HEADER_FILL = PatternFill(
    start_color="2563EB",
    end_color="2563EB",
    fill_type="solid",
)

SECTION_FILL = PatternFill(
    start_color="E0E7FF",
    end_color="E0E7FF",
    fill_type="solid",
)

CORRECT_ANSWER_FILL = PatternFill(
    start_color="DCFCE7",
    end_color="DCFCE7",
    fill_type="solid",
)

HEADER_FONT = Font(
    bold=True,
    size=11,
    color="FFFFFF",
)

SECTION_FONT = Font(
    bold=True,
    color="1E3A8A",
)

THIN_BORDER = Border(
    left=Side(
        style="thin",
        color="D1D5DB",
    ),
    right=Side(
        style="thin",
        color="D1D5DB",
    ),
    top=Side(
        style="thin",
        color="D1D5DB",
    ),
    bottom=Side(
        style="thin",
        color="D1D5DB",
    ),
)

WRAP_ALIGNMENT = Alignment(
    wrap_text=True,
    vertical="top",
)

CENTER_ALIGNMENT = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True,
)


def safe_excel_value(
    value: Any,
) -> Any:
    """
    Подготавливает значение для записи в Excel.

    Excel не поддерживает datetime с часовым поясом.
    Поэтому timezone-aware datetime преобразуется
    в UTC, после чего информация о часовом поясе
    удаляется.

    Строки, похожие на формулы Excel,
    экранируются одинарной кавычкой.
    """
    if value is None:
        return ""

    if isinstance(value, datetime):
        if value.tzinfo is not None:
            value = value.astimezone(
                timezone.utc
            )

            value = value.replace(
                tzinfo=None
            )

        return value

    if isinstance(
        value,
        (
            int,
            float,
            bool,
        ),
    ):
        return value

    enum_value = getattr(
        value,
        "value",
        None,
    )

    if enum_value is not None:
        value = enum_value

    text = str(value)

    if text.startswith(
        (
            "=",
            "+",
            "-",
            "@",
        )
    ):
        return f"'{text}"

    return text


def get_user_name(
    user: Any,
) -> str:
    """
    Возвращает отображаемое имя пользователя.
    """
    if user is None:
        return ""

    full_name = getattr(
        user,
        "full_name",
        None,
    )

    if full_name:
        return str(full_name)

    username = getattr(
        user,
        "username",
        None,
    )

    if username:
        return str(username)

    user_id = getattr(
        user,
        "id",
        None,
    )

    if user_id is not None:
        return f"Пользователь #{user_id}"

    return ""


def get_subject_title(
    variant: Any,
) -> str:
    """
    Возвращает название предмета варианта.
    """
    subject = getattr(
        variant,
        "subject",
        None,
    )

    if subject is not None:
        title = getattr(
            subject,
            "title",
            None,
        )

        if title:
            return str(title)

    subject_id = getattr(
        variant,
        "subject_id",
        None,
    )

    if subject_id is not None:
        return f"Предмет #{subject_id}"

    return ""


def get_status_value(
    value: Any,
) -> str:
    """
    Возвращает строковое значение Enum или строки.
    """
    if value is None:
        return ""

    enum_value = getattr(
        value,
        "value",
        None,
    )

    if enum_value is not None:
        return str(enum_value)

    return str(value)


def get_objective_data(
    question: Any,
) -> tuple[str, str, str]:
    """
    Возвращает код и названия ОРО.

    Если relationship не загружен,
    сохраняется ID ОРО.
    """
    objective = getattr(
        question,
        "learning_objective",
        None,
    )

    if objective is not None:
        return (
            str(
                getattr(
                    objective,
                    "code",
                    "",
                )
                or ""
            ),
            str(
                getattr(
                    objective,
                    "title_kz",
                    "",
                )
                or ""
            ),
            str(
                getattr(
                    objective,
                    "title_ru",
                    "",
                )
                or ""
            ),
        )

    objective_id = getattr(
        question,
        "learning_objective_id",
        None,
    )

    if objective_id is None:
        return "", "", ""

    return (
        f"ID {objective_id}",
        "",
        "",
    )


def get_bilingual_options(
    question: Any,
) -> list[dict[str, Any]]:
    """
    Нормализует варианты ответа вопроса.
    """
    raw_options = getattr(
        question,
        "options",
        None,
    ) or []

    if not isinstance(
        raw_options,
        list,
    ):
        return []

    normalized_options: list[
        dict[str, Any]
    ] = []

    for index, option in enumerate(
        raw_options,
    ):
        if not isinstance(
            option,
            dict,
        ):
            continue

        default_key = chr(
            65 + index
        )

        normalized_options.append(
            {
                "key": str(
                    option.get(
                        "key",
                        default_key,
                    )
                    or default_key
                ),
                "text_kz": str(
                    option.get(
                        "text_kz",
                        option.get(
                            "text",
                            "",
                        ),
                    )
                    or ""
                ),
                "text_ru": str(
                    option.get(
                        "text_ru",
                        option.get(
                            "text",
                            "",
                        ),
                    )
                    or ""
                ),
                "is_correct": (
                    option.get(
                        "is_correct"
                    )
                    is True
                ),
            }
        )

    return normalized_options


def format_options(
    options: list[dict[str, Any]],
    language: str,
) -> str:
    """
    Формирует многострочное представление ответов.
    """
    text_field = (
        "text_kz"
        if language == "kz"
        else "text_ru"
    )

    lines: list[str] = []

    for option in options:
        option_key = option["key"]
        option_text = option[text_field]

        lines.append(
            f"{option_key}. {option_text}"
        )

    return "\n".join(lines)


def get_correct_answer(
    options: list[dict[str, Any]],
) -> str:
    """
    Возвращает обозначение правильного ответа.
    """
    correct_keys = [
        option["key"]
        for option in options
        if option["is_correct"]
    ]

    return ", ".join(
        correct_keys
    )


def get_media_filenames(
    question: Any,
) -> str:
    """
    Возвращает имена изображений вопроса.
    """
    media_files = getattr(
        question,
        "media_files",
        None,
    ) or []

    filenames: list[str] = []

    for media in media_files:
        filename = getattr(
            media,
            "original_filename",
            None,
        )

        if filename:
            filenames.append(
                str(filename)
            )

    return "\n".join(
        filenames
    )


def configure_sheet(
    worksheet: Worksheet,
    headers: list[str],
    widths: list[int],
) -> None:
    """
    Настраивает заголовки, фильтры,
    размеры и закрепление строки.
    """
    for column_index, header in enumerate(
        headers,
        start=1,
    ):
        cell = worksheet.cell(
            row=1,
            column=column_index,
            value=header,
        )

        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = (
            CENTER_ALIGNMENT
        )
        cell.border = THIN_BORDER

    worksheet.freeze_panes = "A2"

    worksheet.auto_filter.ref = (
        f"A1:"
        f"{get_column_letter(len(headers))}"
        f"1"
    )

    worksheet.row_dimensions[
        1
    ].height = 35

    for column_index, width in enumerate(
        widths,
        start=1,
    ):
        worksheet.column_dimensions[
            get_column_letter(
                column_index
            )
        ].width = width


def write_row(
    worksheet: Worksheet,
    row_index: int,
    values: list[Any],
) -> None:
    """
    Записывает одну оформленную строку.
    """
    for column_index, value in enumerate(
        values,
        start=1,
    ):
        cell = worksheet.cell(
            row=row_index,
            column=column_index,
            value=safe_excel_value(
                value
            ),
        )

        cell.border = THIN_BORDER
        cell.alignment = (
            WRAP_ALIGNMENT
        )


def save_workbook(
    workbook: Workbook,
) -> bytes:
    """
    Сохраняет Workbook в bytes.
    """
    buffer = io.BytesIO()

    workbook.save(
        buffer
    )

    buffer.seek(0)

    return buffer.getvalue()


def generate_registry_xlsx(
    questions: list,
) -> bytes:
    """
    Создаёт legacy-реестр отдельных вопросов.

    Функция сохранена для совместимости
    со старым POST /api/v1/export/zip.
    """
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = (
        "Реестр вопросов"
    )

    headers = [
        "ID",
        "Предмет",
        "Тема",
        "Текст вопроса",
        "Вариант A",
        "Вариант B",
        "Вариант C",
        "Вариант D",
        "Правильный ответ",
        "Сложность",
        "Автор",
        "Статус",
    ]

    widths = [
        8,
        18,
        22,
        50,
        30,
        30,
        30,
        30,
        18,
        12,
        24,
        18,
    ]

    configure_sheet(
        worksheet,
        headers,
        widths,
    )

    for row_index, question in enumerate(
        questions,
        start=2,
    ):
        raw_options = getattr(
            question,
            "options",
            None,
        ) or []

        legacy_options = (
            raw_options
            if isinstance(
                raw_options,
                list,
            )
            else []
        )

        correct_letters: list[str] = []

        for option_index, option in enumerate(
            legacy_options,
        ):
            if (
                isinstance(
                    option,
                    dict,
                )
                and option.get(
                    "is_correct"
                )
                is True
            ):
                correct_letters.append(
                    chr(
                        65 + option_index
                    )
                )

        option_texts: list[str] = []

        for option_index in range(4):
            if (
                option_index
                < len(legacy_options)
                and isinstance(
                    legacy_options[
                        option_index
                    ],
                    dict,
                )
            ):
                option = legacy_options[
                    option_index
                ]

                option_texts.append(
                    str(
                        option.get(
                            "text",
                            option.get(
                                "text_ru",
                                "",
                            ),
                        )
                        or ""
                    )
                )
            else:
                option_texts.append("")

        author = getattr(
            question,
            "author",
            None,
        )

        row_data = [
            getattr(
                question,
                "id",
                "",
            ),
            getattr(
                question,
                "subject",
                "",
            )
            or "",
            getattr(
                question,
                "topic",
                "",
            )
            or "",
            getattr(
                question,
                "body",
                "",
            )
            or "",
            *option_texts,
            ", ".join(
                correct_letters
            ),
            getattr(
                question,
                "difficulty",
                "",
            ),
            get_user_name(
                author
            ),
            get_status_value(
                getattr(
                    question,
                    "status",
                    None,
                )
            ),
        ]

        write_row(
            worksheet,
            row_index,
            row_data,
        )

    return save_workbook(
        workbook
    )


def generate_variants_registry_xlsx(
    variants: list,
) -> bytes:
    """
    Создаёт реестр выбранных вариантов.

    Лист «Варианты» содержит сведения
    о вариантах и участниках процесса.

    Лист «Вопросы» содержит полный
    двуязычный состав вопросов.
    """
    workbook = Workbook()

    variants_sheet = workbook.active
    variants_sheet.title = "Варианты"

    questions_sheet = workbook.create_sheet(
        title="Вопросы"
    )

    variant_headers = [
        "ID варианта",
        "Название",
        "Описание",
        "Предмет",
        "Разработчик",
        "Верификатор",
        "Рецензия верификатора",
        "Куратор",
        "Комментарий куратора",
        "Статус",
        "Количество вопросов",
        "Создан",
        "Отправлен",
        "Проверен",
        "Утверждён",
        "Опубликован",
    ]

    variant_widths = [
        14,
        35,
        45,
        24,
        26,
        26,
        50,
        26,
        50,
        18,
        18,
        20,
        20,
        20,
        20,
        20,
    ]

    configure_sheet(
        variants_sheet,
        variant_headers,
        variant_widths,
    )

    question_headers = [
        "ID варианта",
        "Название варианта",
        "Предмет",
        "№ вопроса",
        "ID вопроса",
        "Код ОРО",
        "ОРО KZ",
        "ОРО RU",
        "Когнитивный уровень",
        "Ресурс KZ",
        "Ресурс RU",
        "Текст вопроса KZ",
        "Текст вопроса RU",
        "Варианты ответа KZ",
        "Варианты ответа RU",
        "Правильный ответ",
        "Пояснение KZ",
        "Пояснение RU",
        "Изображения",
        "Статус вопроса",
    ]

    question_widths = [
        14,
        30,
        22,
        12,
        12,
        16,
        35,
        35,
        24,
        45,
        45,
        50,
        50,
        45,
        45,
        18,
        45,
        45,
        35,
        18,
    ]

    configure_sheet(
        questions_sheet,
        question_headers,
        question_widths,
    )

    question_row_index = 2

    for variant_row_index, variant in enumerate(
        variants,
        start=2,
    ):
        questions = list(
            getattr(
                variant,
                "questions",
                None,
            )
            or []
        )

        questions.sort(
            key=lambda question: (
                getattr(
                    question,
                    "order_number",
                    0,
                )
                or 0,
                getattr(
                    question,
                    "id",
                    0,
                )
                or 0,
            )
        )

        write_row(
            variants_sheet,
            variant_row_index,
            [
                getattr(
                    variant,
                    "id",
                    "",
                ),
                getattr(
                    variant,
                    "title",
                    "",
                ),
                getattr(
                    variant,
                    "description",
                    "",
                )
                or "",
                get_subject_title(
                    variant
                ),
                get_user_name(
                    getattr(
                        variant,
                        "developer",
                        None,
                    )
                ),
                get_user_name(
                    getattr(
                        variant,
                        "reviewer",
                        None,
                    )
                ),
                getattr(
                    variant,
                    "review_comment",
                    "",
                )
                or "",
                get_user_name(
                    getattr(
                        variant,
                        "curator",
                        None,
                    )
                ),
                getattr(
                    variant,
                    "curator_comment",
                    "",
                )
                or "",
                get_status_value(
                    getattr(
                        variant,
                        "status",
                        None,
                    )
                ),
                len(questions),
                getattr(
                    variant,
                    "created_at",
                    None,
                ),
                getattr(
                    variant,
                    "submitted_at",
                    None,
                ),
                getattr(
                    variant,
                    "reviewed_at",
                    None,
                ),
                getattr(
                    variant,
                    "approved_at",
                    None,
                ),
                getattr(
                    variant,
                    "published_at",
                    None,
                ),
            ],
        )

        for question in questions:
            objective_code, objective_kz, objective_ru = (
                get_objective_data(
                    question
                )
            )

            options = (
                get_bilingual_options(
                    question
                )
            )

            write_row(
                questions_sheet,
                question_row_index,
                [
                    getattr(
                        variant,
                        "id",
                        "",
                    ),
                    getattr(
                        variant,
                        "title",
                        "",
                    ),
                    get_subject_title(
                        variant
                    ),
                    getattr(
                        question,
                        "order_number",
                        "",
                    ),
                    getattr(
                        question,
                        "id",
                        "",
                    ),
                    objective_code,
                    objective_kz,
                    objective_ru,
                    get_status_value(
                        getattr(
                            question,
                            "cognitive_level",
                            None,
                        )
                    ),
                    getattr(
                        question,
                        "resource_kz",
                        "",
                    )
                    or "",
                    getattr(
                        question,
                        "resource_ru",
                        "",
                    )
                    or "",
                    getattr(
                        question,
                        "question_text_kz",
                        "",
                    )
                    or "",
                    getattr(
                        question,
                        "question_text_ru",
                        "",
                    )
                    or "",
                    format_options(
                        options,
                        "kz",
                    ),
                    format_options(
                        options,
                        "ru",
                    ),
                    get_correct_answer(
                        options
                    ),
                    getattr(
                        question,
                        "explanation_kz",
                        "",
                    )
                    or "",
                    getattr(
                        question,
                        "explanation_ru",
                        "",
                    )
                    or "",
                    get_media_filenames(
                        question
                    ),
                    get_status_value(
                        getattr(
                            question,
                            "status",
                            None,
                        )
                    ),
                ],
            )

            correct_answer_cell = (
                questions_sheet.cell(
                    row=question_row_index,
                    column=16,
                )
            )

            correct_answer_cell.fill = (
                CORRECT_ANSWER_FILL
            )

            question_row_index += 1

    variants_sheet.auto_filter.ref = (
        f"A1:"
        f"{get_column_letter(len(variant_headers))}"
        f"{max(1, variants_sheet.max_row)}"
    )

    questions_sheet.auto_filter.ref = (
        f"A1:"
        f"{get_column_letter(len(question_headers))}"
        f"{max(1, questions_sheet.max_row)}"
    )

    return save_workbook(
        workbook
    )